#!/usr/bin/env python
import yaml
import redis
import requests
import openpyxl

import json
import logging
from datetime import datetime, timedelta

import config

"""Initialize logger."""
log = logging.getLogger()

log.setLevel(logging.INFO)
log.addHandler(logging.StreamHandler())


class Origin():
    def __init__(self, content):
        """Origin representing bus stop.

        Attributes:
            content (dict): Dictionary from YAML parser containing origin data.
        """
        self.code = content["code"]
        self.name = content["name"]

        self.longitude = content["longitude"]
        self.latitude = content["latitude"]

        self.connections = content["connections"]

    def store(self, database):
        """Saves origin data (code, name, position) in Redis.
        Using a custom database schema.

        Args:
            database (obj): Redis database object.
        """
        database.set("origin:{}:name".format(self.code), self.name)
        database.geoadd("origins:positions", self.longitude, self.latitude, self.code)


class Route():
    def __init__(self, origin, destination, duration, timetable):
        """Route between two bus stops. One way.

        Attributes:
            origin (str): Code of route origin.
            destination (str): Code of route destination.
            duration (int): Duration of a journey in minutes.

            timetable (dict): Dictionary from DocumentParser containing timetable.
        """
        self.origin = origin
        self.duration = duration
        self.destination = destination

        self.timetable = timetable

    def store(self, database, expire=True):
        """Saves route data (days, duration) in Redis.
        Using a custom database schema.

        Args:
            database (obj): Redis database object.
            expire (bool): Sets expiration on the key one day after.
        """
        root_key = ("origin:{}:destination:{}".format(self.origin, self.destination))

        database.set(root_key + ":duration", self.duration)
        log.debug("Set duration to: {}".format(self.duration))

        for days in self.timetable:
            log.debug("Added {} day with {} departures.".format(days, len(days)))

            key = root_key + ":date:{}".format(days)
            value = json.dumps(self.timetable[days])
            expiration = datetime.strptime(days, "%d.%m.%Y") + timedelta(days=1)

            database.set(key, value)

            if expire:
                database.expireat(key, expiration)


class DocumentParser():
    def __init__(self, path, sheet=0, start_row=3):
        """Exel document containing the whole timetable.

        Attributes:
            path (str): Path to .xlsx file on local filesystem.

            sheet (int, optional): Number of document sheet.
            start_row (int, optional): Start row of tables.
        """
        self.path = path

        self.workbook = openpyxl.load_workbook(self.path)
        self.sheet = self.workbook[self.workbook.sheetnames[sheet]]

        self.start_row = start_row

    def column(self, column):
        """Parses column for days with departures.

        Args:
            column (int): Number of wanted column.

        Returns:
            A dictionary containing lists with times of departure from selected column.
        """
        departures = {}

        for row in range(self.start_row, self.sheet.max_row + 1):
            date_cell = self.sheet.cell(row=row, column=1)
            data_cell = self.sheet.cell(row=row, column=column)

            if date_cell.value is not None:
                date = date_cell.value.strftime("%d.%m.%Y")
                departures[date] = []

            if data_cell.value is None:
                continue
            else:
                departures[date].append(data_cell.value.strftime("%H:%M"))

        return departures


def download_file(url, filename):
    """Downloads and saves file from URL."""
    with open(filename, "wb") as file:
        file.write(requests.get(url).content)


def parse_yaml(filename):
    """Opens and parser YAML config."""
    with open(filename, "r", encoding="UTF-8") as file:
        return yaml.safe_load(file)["origins"]


def main():
    conf = config.Config()
    log.info("Initialized parsing job for {} month.".format(conf.month))

    log.info("Connecting to Redis database...")
    database = redis.Redis(host=conf.redis_hostname, port=6379, decode_responses=True)

    try:
        url = database.get("url:{}".format(conf.month))
    except:
        log.critical("URL for {} month not found! Exiting...".format(conf.month))
        return

    download_file(url, conf.filename)
    log.warning("Success! File downloaded and saved.")

    parser = DocumentParser(conf.filename)
    log.info("Parsing timetable...")

    yaml_content = parse_yaml(conf.yaml_path)

    log.info("Creating origins...")
    origins = []
    for item in yaml_content:
        origins.append(Origin(item))

    log.info("Creating routes...")
    routes = []
    for origin in origins:
        for route in origin.connections:
            routes.append(Route(origin.code, route["destination"], route["duration"], parser.column(route["column"])))

    log.info("Storing origins data in Redis...")
    for origin in origins:
        origin.store(database)

    log.info("Storing routes data in Redis...")
    for route in routes:
        route.store(database, conf.expire)

    log.warning("Success! Exiting!")


if __name__ == "__main__":
    main()
